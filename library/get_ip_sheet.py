# --- import ---
import common
import sys
import requests
from requests.packages.urllib3.exceptions import InsecureRequestWarning
requests.packages.urllib3.disable_warnings(InsecureRequestWarning)
import json
import openpyxl
from openpyxl.styles.alignment import Alignment
import datetime
from netaddr import *
now = datetime.datetime.now()
now.strftime('%Y-%m-%d %H:%M:%S')

l = common.blueprint()
token = l[0]
bp_id = l[1]
bp_cabling_map_get = common.bp_cabling_map_get(token, bp_id)
bp_racks_get = common.bp_racks_get(token, bp_id)
system_agents_get = common.system_agents_get(token)
bp_qe_post_system_interface = common.bp_qe_post_system_interface(token, bp_id)
bp_qe_post_sec_vn = common.bp_qe_post_sec_vn(token, bp_id)
bp_qe_post_system_interfacemap = common.bp_qe_post_system_interfacemap(token, bp_id)
bp_qe_post_system_vni_vn_vni_interface = common.bp_qe_post_system_vni_vn_vni_interface(token, bp_id)
bp_qe_post_system = common.bp_qe_post_system(token, bp_id)
bp_qe_post_vn = common.bp_qe_post_vn(token, bp_id)



# get [system id, hardware] list
def graph_system_interfacemap():
  d = {}
  for i in bp_qe_post_system_interfacemap['items']:
    d[i['system']['system_id']] = i['map']['device_profile_id']
#  print (json.dumps(d, indent=4))
  return d

# modify xlsx file
def config_xlsx(data, ll, sheetname, ws):
  ws['A1'] = sheetname
  ws['A1'].font = openpyxl.styles.Font(bold=True, size=16)
  ws['A2'] = ''
  ws['A3'] = now
  ws['A3'].alignment = Alignment(horizontal = 'left')
  ws['A4'] = ''
  ws.append(data)
  # sort column
  if sheetname != 'IP':
    for i in sorted(ll):
      ws.append(i)
  else:
    j = 6
    for i in ll:
      ws['C' + str(j)] = i
      j = j + 1
  # adjust column dimension
  max_len = 0
  for i in ws.columns:
    column = i[0].coordinate[:-1]
    for j in i:
      if len(str(j.value)) > max_len:
        max_len = len(str(j.value))
    adj_width = (max_len) * 1.1
    ws.column_dimensions[column].width = adj_width
    max_len = 0


class ConfigSheet:

  def __init__(self):
    self.list = ['Device','Cabling','Underlay','Overlay','IP']

  def make_sheet_device(self):
    l = [] # tmp list
    ll = [] # [['', 'spine', 'spine1', '2CC260993101', '192.168.59.31'], ['', 'spine', 'spine2', '2CC260993201', '192.168.59.32']]
    d = {} # {'2CC260994301': ['192.168.59.43', '4.21.5F'], '2CC260994201': ['192.168.59.42', '3.7.5'], '2CC260993201'
    dd = {} # leaf IDs in list - {'rack_type_1_001': ['dc1b20f2-022c-4b5f-a371-14bbc47c64ab', '6d30e4ce-c073-4e76-a422-82d47c4438f8'], 'rack_type_2_001': ['36db970f-0241-48a8-87c9-1f99a4ee8cb7', '89ad07cd-06e0-4ba2-aaa2-6c50aabd8e96']}
    # first line for 'Device'
    data = ['Rack', 'Role', 'Hostname', 'Management IP', 'Hardware', 'System ID', 'NOS Version']
    # make d
    for i in system_agents_get['items']:
      if i['status']['operation_mode'] == "full_control":
        l.extend([i['running_config']['management_ip'], i['status']['platform_version']])
        d[i['status']['system_id']] = l
        l = []
    # make dd
    for i in bp_racks_get['items']:
      dd[i['label']] = [j['id'] for j in i['leafs'] ]
      l = []
    # make contents for xlsx
    for i in bp_qe_post_system['items']:
      i = i['system']
      # spine
      if i['role'] == 'spine':
        # when BP Physical Devices already set
        if i['system_id'] != None: l.extend(['', i['role'], i['hostname'], d[i['system_id']][0], graph_system_interfacemap()[i['system_id']] ,i['system_id'], d[i['system_id']][1]])
        # not set duo to demo or else
        else: l.extend(['', i['role'], i['hostname'], '', '', '', ''])
        ll.append(l)
        l = []
      # leaf
      for j in dd.items():
        if i['id'] in j[1]:
          # When BP Physical Devices already set
          if i['system_id'] != None: l.extend([j[0], i['role'], i['hostname'], d[i['system_id']][0], graph_system_interfacemap()[i['system_id']], i['system_id'], d[i['system_id']][1]])
          # not set duo to demo or else
          else: l.extend([j[0], i['role'], i['hostname'], '', '', '', ''])
          ll.append(l)
          l = []
    # replace None to '' for sort
    ll = [['' if i is None else i for i in j] for j in ll]
    # create xlsx file
    wb = openpyxl.Workbook()
    ws = wb.create_sheet(self.list[0])
    config_xlsx(data, ll, self.list[0], ws)
    # Remove 'Sheet'
    wb.remove(wb.worksheets[0])
    wb.save('ip_sheet.xlsx')


  def make_sheet_cabling_map(self):
    l = []
    ll = []
    d = {}
    data = ['Rack Label', 'System ID 1', 'Role', 'System Label 1', 'Interafce', 'IPv4 Address', 'System ID 2', 'Role', 'System Label 2', 'Interafce', 'IPv4 Address', 'Speed']
    # make rack dict. {'rack_type_1_001': ['0de8be6f-ab11-41a3-9a7d-c1d95cba5e2a', '54337a2f-f913-4fbe-96eb-f383bd91e22d',
    for i in bp_racks_get['items']:
      d[i['label']] = [j['id'] for j in i['servers']] + [j['id'] for j in i['leafs']]
    # make cabling list
    for i in bp_cabling_map_get['links']:
      for j in i['endpoints']:
        if j['system']['role'] == 'leaf' or j['system']['role'] == 'spine': l.extend([j['system']['role'], j['system']['label'], j['interface']['if_name'], j['interface']['ipv4_addr']])
        elif j['system']['role'] == 'external_router': l.extend([j['system']['role'], j['system']['label'], '', j['interface']['ipv4_addr']])
        else: l.extend([j['system']['role'], j['system']['label'], '', ''])
        # add a leaf id in front of system label for checking at rack and interfacemap.
        l.insert(-4, j['system']['id'])
      l.append(i['speed'])
      # add rack label in front of cabling list.
      for j in d.items(): # j:('rack_type_2_001', ['fb82532f-6831-4807-a908-8208cf47f2a2', '1084e191-1cad-4164-8c6c-ade962f5647e', '36db970f-0241-48a8-87c9-1f99a4ee8cb7', '89ad07cd-06e0-4ba2-aaa2-6c50aabd8e96'])
        if l[0] in j[1] or l[5] in j[1]: l.insert(0, j[0])
      # replace system id to serial number.
      for j in bp_qe_post_system['items']:
        if l[1] == j['system']['id']: l[1] = j['system']['system_id']
        elif l[6] == j['system']['id']: l[6] = j['system']['system_id']
      ll.append(l)
      l = []
    ll = [['' if i is None else i for i in j] for j in ll]
    wb = openpyxl.load_workbook('ip_sheet.xlsx')
    ws = wb.create_sheet(self.list[1])
    config_xlsx(data, ll, self.list[1], ws)
    wb.save('ip_sheet.xlsx')


  def make_sheet_underlay(self):
    l = []
    ll = []
    # First line for 'Underlay'
    data = ['Hostname', 'Int Type', 'Int Name', 'IPv4 Address', 'Protocol']
    for i in bp_qe_post_system_interface['items']:
      if i['system']['hostname'] != None and i['interface']['ipv4_addr'] != None:
        l.extend([i['system']['hostname'], i['interface']['if_type'], i['interface']['if_name'], i['interface']['ipv4_addr'], i['interface']['protocols']])
        ll.append(l)
        l = []
    ll = [['' if i is None else i for i in j] for j in ll]
    wb = openpyxl.load_workbook('ip_sheet.xlsx')
    ws = wb.create_sheet(self.list[2])
    config_xlsx(data, ll, self.list[2], ws)
    wb.save('ip_sheet.xlsx')


  def make_sheet_overlay(self):
    l = []
    ll = []
    data = ['Hostname', 'VRF_Name', 'VRF_VLAN', 'VRF_VNI', 'Int_type', 'VN_Name', 'VN_VLAN', 'VN_VXLAN', 'NW_Address', 'IP_Address', 'IP_Gateway']
    for i in bp_qe_post_system_vni_vn_vni_interface['items']:
      for j in bp_qe_post_sec_vn['items']:
        if i['virtual_network']['id'] == j['virtual_network']['id']:
          l.extend([i['system']['hostname'], j['vrf']['vrf_name'], j['vrf']['vlan_id'], j['vrf']['vni_id'], i['interface']['if_type'],i['virtual_network']['label'],i['vn_instance']['vlan_id'],i['virtual_network']['vn_id'],i['virtual_network']['ipv4_subnet'],i['interface']['ipv4_addr'],i['virtual_network']['virtual_gateway_ipv4']])
          ll.append(l)
          l = []
    wb = openpyxl.load_workbook('ip_sheet.xlsx')
    ws = wb.create_sheet(self.list[3])
    config_xlsx(data, ll, self.list[3], ws)
    wb.save('ip_sheet.xlsx')



ConfigSheet().make_sheet_device()
ConfigSheet().make_sheet_cabling_map()
ConfigSheet().make_sheet_underlay()
ConfigSheet().make_sheet_overlay()
